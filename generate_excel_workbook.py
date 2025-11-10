#!/usr/bin/env python3
"""
Generate BookWise Excel Workbook
Creates a comprehensive Excel workbook from existing CSV data
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

# Paths
OUTPUT_DIR = Path('output')
EXCEL_FILE = OUTPUT_DIR / 'bookwise_library.xlsx'

# Color scheme
COLORS = {
    'header': '4472C4',
    'legendary': '00B050',
    'exceptional': '92D050',
    'outstanding': 'FFFF00',
    'excellent': 'FFC000',
    'good': 'FF9999',
    'warning': 'FFC7CE',
}


def load_data():
    """Load all necessary data"""
    print("Loading data...")

    # Load main scored books
    scores_df = pd.read_csv(OUTPUT_DIR / 'all_books_quality_scores.csv')

    # Load owned books from JSONL (new format)
    import json
    owned_books = {}
    jsonl_file = OUTPUT_DIR / 'calibre_enrichment_ready.jsonl'

    if jsonl_file.exists():
        with open(jsonl_file, 'r', encoding='utf-8') as f:
            for line in f:
                book = json.loads(line)
                # Create key (title, author) and store calibre_id
                key = (book['title'], book['author'])
                owned_books[key] = book['calibre_id']

    owned_ids = set(owned_books.keys())

    # Add owned column
    scores_df['owned'] = scores_df.apply(
        lambda row: 'yes' if (row['title'], row['author']) in owned_ids else 'no',
        axis=1
    )

    # Add calibre_id from JSONL
    scores_df['calibre_id'] = scores_df.apply(
        lambda row: owned_books.get((row['title'], row['author']), ''),
        axis=1
    )

    # Add quality tier
    def get_tier(score):
        if score >= 90: return 'Legendary'
        if score >= 80: return 'Exceptional'
        if score >= 70: return 'Outstanding'
        if score >= 60: return 'Excellent'
        if score >= 50: return 'Very Good'
        if score >= 40: return 'Good'
        if score >= 30: return 'Solid'
        if score >= 20: return 'Notable'
        if score >= 10: return 'Recognized'
        return 'Limited'

    scores_df['quality_tier'] = scores_df['final_score'].apply(get_tier)

    print(f"  Loaded {len(scores_df)} books")
    print(f"  {len(owned_ids)} books owned")

    return scores_df


def create_workbook():
    """Create the Excel workbook"""
    print("\nCreating Excel workbook...")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Load data
    df = load_data()

    # Create sheets
    create_dashboard(wb, df)
    create_all_books(wb, df)
    create_owned_books(wb, df)
    create_missing_books(wb, df)
    create_shopping_list(wb, df)
    create_statistics(wb, df)
    create_pivot_analysis(wb, df)
    create_import_sheets(wb, df)

    # Save
    wb.save(EXCEL_FILE)
    print(f"\n✅ Excel workbook created: {EXCEL_FILE}")
    print(f"   File size: {EXCEL_FILE.stat().st_size / 1024 / 1024:.2f} MB")


def create_dashboard(wb, df):
    """Create Dashboard sheet"""
    print("  Creating Dashboard sheet...")
    ws = wb.create_sheet("Dashboard", 0)

    # Title
    ws['A1'] = "📊 BookWise Library Analysis"
    ws['A1'].font = Font(size=18, bold=True, color='1F4E78')
    ws.merge_cells('A1:E1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)

    # Key Metrics
    ws['A4'] = "Key Metrics"
    ws['A4'].font = Font(size=14, bold=True)

    owned_count = len(df[df['owned'] == 'yes'])
    total_count = len(df)

    metrics = [
        ('Total Scored Books', total_count),
        ('Books You Own', owned_count),
        ('Coverage %', f"{owned_count/total_count*100:.1f}%"),
        ('Average Score (Owned)', f"{df[df['owned']=='yes']['final_score'].mean():.1f}"),
        ('Legendary Books (90+) Owned', len(df[(df['owned']=='yes') & (df['final_score']>=90)])),
        ('Books Worth Buying', total_count - owned_count),
    ]

    for i, (label, value) in enumerate(metrics, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)

        # Add border
        for cell in [ws[f'A{i}'], ws[f'B{i}']]:
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Quick Navigation
    ws['A13'] = "🔗 Quick Navigation"
    ws['A13'].font = Font(size=12, bold=True)

    links = [
        ('View All Books', 'All Books'),
        ('See Shopping List', 'Shopping List'),
        ('Browse Owned Books', 'Owned Books'),
        ('Import to Calibre', 'Import_Metadata'),
    ]

    for i, (text, sheet) in enumerate(links, start=14):
        ws[f'A{i}'] = text
        ws[f'A{i}'].hyperlink = f"#{sheet}!A1"
        ws[f'A{i}'].font = Font(color='0563C1', underline='single')

    # Score distribution data
    ws['D4'] = "Score Distribution"
    ws['D4'].font = Font(size=12, bold=True)

    score_ranges = [
        ('90+ (Legendary)', len(df[df['final_score'] >= 90])),
        ('80-89 (Exceptional)', len(df[(df['final_score'] >= 80) & (df['final_score'] < 90)])),
        ('70-79 (Outstanding)', len(df[(df['final_score'] >= 70) & (df['final_score'] < 80)])),
        ('60-69 (Excellent)', len(df[(df['final_score'] >= 60) & (df['final_score'] < 70)])),
        ('50-59 (Very Good)', len(df[(df['final_score'] >= 50) & (df['final_score'] < 60)])),
        ('<50', len(df[df['final_score'] < 50])),
    ]

    for i, (label, count) in enumerate(score_ranges, start=5):
        ws[f'D{i}'] = label
        ws[f'E{i}'] = count

    # Add chart
    chart = BarChart()
    chart.title = "Score Distribution"
    chart.x_axis.title = "Score Range"
    chart.y_axis.title = "Number of Books"

    data = Reference(ws, min_col=5, min_row=5, max_row=10)
    cats = Reference(ws, min_col=4, min_row=5, max_row=10)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20

    ws.add_chart(chart, "G4")

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12


def create_all_books(wb, df):
    """Create All Books sheet"""
    print("  Creating All Books sheet...")
    ws = wb.create_sheet("All Books")

    # Prepare data
    df_export = df[['title', 'author', 'final_score', 'owned', 'calibre_id',
                    'awards_and_recognition', 'quality_tier']].copy()
    df_export.columns = ['Title', 'Author', 'Score', 'Owned', 'Calibre_ID', 'Recognition', 'Quality_Tier']

    # Write headers
    headers = list(df_export.columns)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Format score column
            if c_idx == 3:  # Score column
                score = value
                if score >= 90:
                    cell.fill = PatternFill(start_color=COLORS['legendary'], end_color=COLORS['legendary'], fill_type='solid')
                    cell.font = Font(bold=True)
                elif score >= 80:
                    cell.fill = PatternFill(start_color=COLORS['exceptional'], end_color=COLORS['exceptional'], fill_type='solid')
                elif score >= 70:
                    cell.fill = PatternFill(start_color=COLORS['outstanding'], end_color=COLORS['outstanding'], fill_type='solid')

            # Format owned column
            if c_idx == 4:  # Owned column
                if value == 'yes':
                    cell.value = '✅ yes'
                else:
                    cell.value = '❌ no'

    # Add data validation for Owned column
    dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f'D2:D{len(df_export)+1}')

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f'A1:G{len(df_export)+1}'

    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 15


def create_owned_books(wb, df):
    """Create Owned Books sheet"""
    print("  Creating Owned Books sheet...")
    ws = wb.create_sheet("Owned Books")

    owned_df = df[df['owned'] == 'yes'].copy()
    owned_df = owned_df.sort_values('final_score', ascending=False)

    # Prepare data
    df_export = owned_df[['title', 'author', 'final_score', 'awards_and_recognition', 'quality_tier']].copy()
    df_export.columns = ['Title', 'Author', 'Score', 'Recognition', 'Quality_Tier']

    # Write headers
    headers = list(df_export.columns)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Summary row
    summary_row = len(df_export) + 3
    ws[f'A{summary_row}'] = "SUMMARY"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'] = f"Total Owned: {len(owned_df)}"
    ws[f'C{summary_row}'] = f"Avg: {owned_df['final_score'].mean():.1f}"

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15


def create_missing_books(wb, df):
    """Create Missing Books sheet"""
    print("  Creating Missing Books sheet...")
    ws = wb.create_sheet("Missing Books")

    missing_df = df[df['owned'] == 'no'].copy()
    missing_df = missing_df.sort_values('final_score', ascending=False)

    # Add priority column
    def get_priority(score):
        if score >= 90: return '🔴 Must Buy'
        if score >= 80: return '🟠 High Priority'
        if score >= 70: return '🟡 Medium Priority'
        if score >= 60: return '🟢 Low Priority'
        return 'Optional'

    missing_df['priority'] = missing_df['final_score'].apply(get_priority)

    # Prepare data
    df_export = missing_df[['title', 'author', 'final_score', 'priority', 'awards_and_recognition']].copy()
    df_export.columns = ['Title', 'Author', 'Score', 'Priority', 'Recognition']

    # Write headers
    headers = list(df_export.columns)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')

    # Write data (limit to first 1000 for performance)
    for r_idx, row in enumerate(dataframe_to_rows(df_export.head(1000), index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50


def create_shopping_list(wb, df):
    """Create Shopping List sheet"""
    print("  Creating Shopping List sheet...")
    ws = wb.create_sheet("Shopping List")

    # Title
    ws['A1'] = "🛒 SHOPPING LIST - Top 50 Books Worth Buying"
    ws['A1'].font = Font(size=16, bold=True, color='1F4E78')
    ws.merge_cells('A1:E1')

    ws['A2'] = "Auto-sorted by score | Only books you don't own"
    ws['A2'].font = Font(size=10, italic=True)

    # Get top 50 missing books
    missing_df = df[df['owned'] == 'no'].copy()
    top_50 = missing_df.nlargest(50, 'final_score')

    # Add why buy column
    def get_why_buy(row):
        score = row['final_score']
        recognition = row['awards_and_recognition']

        if score >= 95:
            return "Highest rated masterpiece"
        elif score >= 90:
            return "Legendary classic"
        elif score >= 85:
            return "Award-winning excellence"
        elif score >= 80:
            return "Critically acclaimed"
        else:
            return "Highly regarded"

    top_50['why_buy'] = top_50.apply(get_why_buy, axis=1)
    top_50['rank'] = range(1, len(top_50) + 1)

    # Prepare data
    df_export = top_50[['rank', 'title', 'author', 'final_score', 'awards_and_recognition', 'why_buy']].copy()
    df_export.columns = ['Rank', 'Title', 'Author', 'Score', 'Awards/Lists', 'Why Buy']

    # Write headers
    headers = list(df_export.columns)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=False), start=5):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Highlight top 10
            if c_idx == 1 and value <= 10:
                cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
                cell.font = Font(bold=True)

    # Set column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 30


def create_statistics(wb, df):
    """Create Statistics sheet"""
    print("  Creating Statistics sheet...")
    ws = wb.create_sheet("Statistics")

    owned_df = df[df['owned'] == 'yes']
    missing_df = df[df['owned'] == 'no']

    # Title
    ws['A1'] = "📈 Library Statistics"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')

    # Overall statistics
    ws['A3'] = "Overall Statistics"
    ws['A3'].font = Font(size=12, bold=True)

    overall_stats = [
        ('Total Scored Books', len(df)),
        ('Books Owned', len(owned_df)),
        ('Books Missing', len(missing_df)),
        ('Coverage %', f"{len(owned_df)/len(df)*100:.1f}%"),
        ('Average Score (All)', f"{df['final_score'].mean():.1f}"),
        ('Average Score (Owned)', f"{owned_df['final_score'].mean():.1f}"),
        ('Median Score (Owned)', f"{owned_df['final_score'].median():.1f}"),
        ('Highest Score', df['final_score'].max()),
        ('Lowest Score', df['final_score'].min()),
    ]

    for i, (label, value) in enumerate(overall_stats, start=4):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)

    # Score distribution
    ws['D3'] = "Score Distribution"
    ws['D3'].font = Font(size=12, bold=True)

    ws['D4'] = "Score Range"
    ws['E4'] = "Count"
    ws['F4'] = "Percent"

    for cell in [ws['D4'], ws['E4'], ws['F4']]:
        cell.font = Font(bold=True)

    score_ranges = [
        ('90+ (Legendary)', len(df[df['final_score'] >= 90])),
        ('80-89 (Exceptional)', len(df[(df['final_score'] >= 80) & (df['final_score'] < 90)])),
        ('70-79 (Outstanding)', len(df[(df['final_score'] >= 70) & (df['final_score'] < 80)])),
        ('60-69 (Excellent)', len(df[(df['final_score'] >= 60) & (df['final_score'] < 70)])),
        ('50-59 (Very Good)', len(df[(df['final_score'] >= 50) & (df['final_score'] < 60)])),
        ('<50', len(df[df['final_score'] < 50])),
    ]

    for i, (label, count) in enumerate(score_ranges, start=5):
        ws[f'D{i}'] = label
        ws[f'E{i}'] = count
        ws[f'F{i}'] = f"{count/len(df)*100:.1f}%"

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12


def create_pivot_analysis(wb, df):
    """Create pivot analysis sheet with various data breakdowns"""
    print("  Creating Pivot Analysis sheet...")
    ws = wb.create_sheet("Pivot Analysis")

    # Title
    ws['A1'] = "📊 Pivot Analysis"
    ws['A1'].font = Font(size=16, bold=True, color='1F4E78')
    ws.merge_cells('A1:F1')

    row = 3

    # ========== PIVOT 1: Books by Author ==========
    ws[f'A{row}'] = "Books by Author (Top 30)"
    ws[f'A{row}'].font = Font(size=13, bold=True, color='1F4E78')
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    # Create pivot data
    author_pivot = df.groupby('author').agg({
        'title': 'count',
        'final_score': ['mean', 'max', 'min'],
        'owned': lambda x: (x == 'yes').sum()
    }).round(1)

    author_pivot.columns = ['Total Books', 'Avg Score', 'Max Score', 'Min Score', 'Owned']
    author_pivot = author_pivot.sort_values('Total Books', ascending=False).head(30)
    author_pivot['Missing'] = author_pivot['Total Books'] - author_pivot['Owned']
    author_pivot = author_pivot[['Total Books', 'Owned', 'Missing', 'Avg Score', 'Max Score', 'Min Score']]

    # Write headers
    headers = ['Author', 'Total Books', 'Owned', 'Missing', 'Avg Score', 'Max Score', 'Min Score']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Write data
    start_row = row
    for author, data in author_pivot.iterrows():
        ws.cell(row=row, column=1, value=author)
        ws.cell(row=row, column=2, value=int(data['Total Books']))
        ws.cell(row=row, column=3, value=int(data['Owned']))
        ws.cell(row=row, column=4, value=int(data['Missing']))
        ws.cell(row=row, column=5, value=float(data['Avg Score']))
        ws.cell(row=row, column=6, value=float(data['Max Score']))
        ws.cell(row=row, column=7, value=float(data['Min Score']))

        # Color code owned percentage
        owned_pct = data['Owned'] / data['Total Books']
        if owned_pct >= 0.8:
            ws.cell(row=row, column=3).fill = PatternFill(start_color=COLORS['legendary'],
                                                          end_color=COLORS['legendary'], fill_type='solid')
        elif owned_pct >= 0.5:
            ws.cell(row=row, column=3).fill = PatternFill(start_color=COLORS['exceptional'],
                                                          end_color=COLORS['exceptional'], fill_type='solid')

        row += 1

    row += 2

    # ========== PIVOT 2: Books by Quality Tier ==========
    ws[f'A{row}'] = "Books by Quality Tier"
    ws[f'A{row}'].font = Font(size=13, bold=True, color='1F4E78')
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    # Create pivot data
    tier_pivot = df.groupby('quality_tier').agg({
        'title': 'count',
        'owned': lambda x: (x == 'yes').sum()
    })
    tier_pivot.columns = ['Total', 'Owned']
    tier_pivot['Missing'] = tier_pivot['Total'] - tier_pivot['Owned']
    tier_pivot['Owned %'] = (tier_pivot['Owned'] / tier_pivot['Total'] * 100).round(1)

    # Sort by quality tier order
    tier_order = ['Legendary', 'Exceptional', 'Outstanding', 'Excellent', 'Very Good',
                  'Good', 'Solid', 'Notable', 'Recognized', 'Limited']
    tier_pivot = tier_pivot.reindex([t for t in tier_order if t in tier_pivot.index])

    # Write headers
    headers = ['Quality Tier', 'Total Books', 'Owned', 'Missing', 'Owned %']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Write data
    for tier, data in tier_pivot.iterrows():
        ws.cell(row=row, column=1, value=tier)
        ws.cell(row=row, column=2, value=int(data['Total']))
        ws.cell(row=row, column=3, value=int(data['Owned']))
        ws.cell(row=row, column=4, value=int(data['Missing']))
        ws.cell(row=row, column=5, value=f"{data['Owned %']:.1f}%")

        # Color code the tier name
        if tier == 'Legendary':
            ws.cell(row=row, column=1).fill = PatternFill(start_color=COLORS['legendary'],
                                                          end_color=COLORS['legendary'], fill_type='solid')
        elif tier == 'Exceptional':
            ws.cell(row=row, column=1).fill = PatternFill(start_color=COLORS['exceptional'],
                                                          end_color=COLORS['exceptional'], fill_type='solid')
        elif tier == 'Outstanding':
            ws.cell(row=row, column=1).fill = PatternFill(start_color=COLORS['outstanding'],
                                                          end_color=COLORS['outstanding'], fill_type='solid')

        row += 1

    row += 2

    # ========== PIVOT 3: Score Distribution by Ownership ==========
    ws[f'A{row}'] = "Score Distribution by Ownership"
    ws[f'A{row}'].font = Font(size=13, bold=True, color='1F4E78')
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    # Create pivot data
    score_bins = [(90, 100, '90-100'), (80, 89, '80-89'), (70, 79, '70-79'),
                  (60, 69, '60-69'), (50, 59, '50-59'), (0, 49, '0-49')]

    # Write headers
    headers = ['Score Range', 'Total Books', 'Owned', 'Missing', 'Owned %']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Write data
    for min_score, max_score, label in score_bins:
        mask = (df['final_score'] >= min_score) & (df['final_score'] <= max_score)
        total = mask.sum()
        owned = ((mask) & (df['owned'] == 'yes')).sum()
        missing = total - owned
        owned_pct = (owned / total * 100) if total > 0 else 0

        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=int(total))
        ws.cell(row=row, column=3, value=int(owned))
        ws.cell(row=row, column=4, value=int(missing))
        ws.cell(row=row, column=5, value=f"{owned_pct:.1f}%")

        row += 1

    row += 2

    # ========== PIVOT 4: Top Authors by Average Score (min 3 books) ==========
    ws[f'A{row}'] = "Top Authors by Average Score (Min 3 Books)"
    ws[f'A{row}'].font = Font(size=13, bold=True, color='1F4E78')
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    # Create pivot data
    author_score_pivot = df.groupby('author').agg({
        'title': 'count',
        'final_score': 'mean',
        'owned': lambda x: (x == 'yes').sum()
    }).round(1)

    author_score_pivot.columns = ['Book Count', 'Avg Score', 'Owned']
    author_score_pivot = author_score_pivot[author_score_pivot['Book Count'] >= 3]
    author_score_pivot = author_score_pivot.sort_values('Avg Score', ascending=False).head(20)

    # Write headers
    headers = ['Author', 'Book Count', 'Avg Score', 'Owned']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Write data
    for author, data in author_score_pivot.iterrows():
        ws.cell(row=row, column=1, value=author)
        ws.cell(row=row, column=2, value=int(data['Book Count']))
        ws.cell(row=row, column=3, value=float(data['Avg Score']))
        ws.cell(row=row, column=4, value=int(data['Owned']))

        # Highlight high average scores
        if data['Avg Score'] >= 85:
            ws.cell(row=row, column=3).fill = PatternFill(start_color=COLORS['legendary'],
                                                          end_color=COLORS['legendary'], fill_type='solid')
            ws.cell(row=row, column=3).font = Font(bold=True)
        elif data['Avg Score'] >= 80:
            ws.cell(row=row, column=3).fill = PatternFill(start_color=COLORS['exceptional'],
                                                          end_color=COLORS['exceptional'], fill_type='solid')

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12


def create_import_sheets(wb, df):
    """Create import sheets for Calibre"""
    print("  Creating Import sheets...")

    # Import Metadata
    ws_meta = wb.create_sheet("Import_Metadata")
    ws_meta['A1'] = "CALIBRE IMPORT - Metadata"
    ws_meta['A1'].font = Font(size=14, bold=True)

    ws_meta['A2'] = "Instructions:"
    ws_meta['A3'] = "1. File > Save As > CSV (Comma delimited)"
    ws_meta['A4'] = "2. Save as: import_metadata.csv"
    ws_meta['A5'] = "3. Run: calibredb set_metadata --fields-from-csv import_metadata.csv"

    # Headers
    headers_meta = ['id', 'title', 'authors', 'tags', 'comments']
    for col, header in enumerate(headers_meta, start=1):
        ws_meta.cell(row=7, column=col, value=header).font = Font(bold=True)

    # Add sample data (owned books only)
    owned_df = df[df['owned'] == 'yes'].head(100)  # Limit for demo
    for r_idx, (_, row) in enumerate(owned_df.iterrows(), start=8):
        ws_meta.cell(row=r_idx, column=1, value='')  # Calibre ID
        ws_meta.cell(row=r_idx, column=2, value=row['title'])
        ws_meta.cell(row=r_idx, column=3, value=row['author'])
        ws_meta.cell(row=r_idx, column=4, value=f"Score: {row['quality_tier']}, Quality: Processed")
        ws_meta.cell(row=r_idx, column=5, value=row['awards_and_recognition'])

    # Import Tags
    ws_tags = wb.create_sheet("Import_Tags")
    ws_tags['A1'] = "CALIBRE IMPORT - Tags Only"
    ws_tags['A1'].font = Font(size=14, bold=True)

    ws_tags['A2'] = "Quick tag import (id, tags columns only)"

    headers_tags = ['id', 'tags']
    for col, header in enumerate(headers_tags, start=1):
        ws_tags.cell(row=4, column=col, value=header).font = Font(bold=True)

    # Add sample data
    for r_idx, (_, row) in enumerate(owned_df.iterrows(), start=5):
        ws_tags.cell(row=r_idx, column=1, value='')  # Calibre ID
        ws_tags.cell(row=r_idx, column=2, value=f"Score: {row['quality_tier']}, Quality: Processed")

    # Set column widths
    for ws in [ws_meta, ws_tags]:
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 50


if __name__ == '__main__':
    print("=" * 60)
    print("BookWise Excel Workbook Generator")
    print("=" * 60)

    create_workbook()

    print("\n" + "=" * 60)
    print("✅ Done! Open the file to explore:")
    print(f"   {EXCEL_FILE.absolute()}")
    print("=" * 60)
